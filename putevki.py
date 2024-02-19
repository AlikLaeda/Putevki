# Путёвки для выпуска автобусов для пассажирских перевозок
# на основе формы 6

# Переделываем базовый шаблон путёвки, шаблон и результат в xlsx.
# 
# Для сбора инвормации используеи форму на tkinter.
# Водителя и автобус выбираем из списка.
#
# Для хранения последней информации и хранения списков водителей
# и автобусов используется json.

from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import ttk
import tkinter.font as tkFont
import json
import time

# Функция сохранения
#
def save_settings():
    # Сохраняем словари с водителями и автобусами а данные из формы в
    # файл. Словари меняются редко, но перезаписываем их каждый раз.
    # Обязательно ensure_ascii=False без этого кирилические символы
    # экранируются.
    data = {
        'dict_of_buses': dict_of_buses,
        'dict_of_drivers': dict_of_drivers,
        'txt_ser':txt_ser.get(),
        'txt_nom':txt_nom.get(),
        'txt_chi':txt_chi.get(),
        'txt_mes':txt_mes.get(),
        'txt_chiu':txt_chiu.get(),
        'txt_mesu':txt_mesu.get(),
        'txt_nac':txt_nac.get(),
        'txt_kon':txt_kon.get(),
        'txt_per':txt_per.get()
        }
    with open("settings.json", "w", encoding="utf-8") as write_file:
        json.dump(data, write_file, ensure_ascii=False, indent='  ')



# Вставляем в поля начала и конца смены, перерыв
# несколько стандартных сценариев подвешенных на кнопки.

def clic_06():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "06:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "22:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 14:00 до 17:00")
    
def clic_07():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "07:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "22:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 13:00 до 16:00")

def clic_22():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "22:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "07:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 01:00 до 04:00")

def clic_b10():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "10:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "22:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 13:00 до 16:00")

def clic_b22():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "22:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "10:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 01:00 до 04:00")



# Копируем дату старта смены в дату окончания смены
def clic_duble():
    txt_chiu.delete(0, END)
    txt_chiu.insert(0,txt_chi.get())
    txt_mesu.delete(0, END)
    txt_mesu.insert(0,txt_mes.get())

# Обрабатываем кнопки + и - номера путёвки
def clic_increment():
    a = str(int(txt_nom.get()) + 1)
    if len(a)==1: a='0'+a
    txt_nom.delete(0, END)
    txt_nom.insert(0,a)

def clic_decrement():
    a = str(int(txt_nom.get()) - 1)
    if len(a)==1: a='0'+a
    txt_nom.delete(0, END)
    txt_nom.insert(0,a)

# Обрабатываем кнопки + и - дат выезда и возвращения
def clic_incdata():
    a = str(int(txt_chi.get()) + 1)
    if len(a)==1: a='0'+a
    txt_chi.delete(0, END)
    txt_chi.insert(0,a)

def clic_decdata():
    a = str(int(txt_chi.get()) - 1)
    if len(a)==1: a='0'+a
    txt_chi.delete(0, END)
    txt_chi.insert(0,a)

def clic_incdatau():
    a = str(int(txt_chiu.get()) + 1)
    if len(a)==1: a='0'+a
    txt_chiu.delete(0, END)
    txt_chiu.insert(0,a)

def clic_decdatau():
    a = str(int(txt_chiu.get()) - 1)
    if len(a)==1: a='0'+a
    txt_chiu.delete(0, END)
    txt_chiu.insert(0,a)


# Обрабатываем кнопку "создать". Изменяем экселевскую таблицу
# берём новые значения из элементов формы. Сохраняем таблицу в
# файл в папке
def clic_create():
    ws['BZ2'] = txt_ser.get()
    ws['CJ2'] = txt_nom.get()
    ws['AN4'] = txt_chi.get()
    ws['AS4'] = txt_mes.get()
    ws['BY4'] = txt_chiu.get()
    ws['CG4'] = txt_mesu.get()

    # Берём данные автобуса из словаря.
    ws['BQ6'] = dict_of_buses.get(combo_buses.get())[0]
    ws['CZ6'] = dict_of_buses.get(combo_buses.get())[1]
    ws['CH7'] = dict_of_buses.get(combo_buses.get())[3]
    ws['CC9'] = dict_of_buses.get(combo_buses.get())[4]
    ws['BY16'] = txt_nac.get()
    ws['BY18'] = txt_kon.get()
    ws['BY20'] = txt_per.get()
    ws['P25'] = combo_drivers.get()
    ws['P51'] = combo_drivers.get()

    # Берём данные водителя из словаря
    ws['BJ30'] = dict_of_drivers.get(combo_drivers.get())[0]
    ws['BJ32'] = dict_of_drivers.get(combo_drivers.get())[1]
    ws['BJ34'] = dict_of_drivers.get(combo_drivers.get())[2]
    ws['CQ30'] = dict_of_drivers.get(combo_drivers.get())[3]
    ws['DC30'] = dict_of_drivers.get(combo_drivers.get())[4]
    ws['DN30'] = dict_of_drivers.get(combo_drivers.get())[5]
    ws['BZ10'] = combo_price.get()
    ws['BQ11'] = combo_contry.get()
    ws.title = '{} {}'.format(combo_drivers.get(), txt_nac.get()[:2])

    # В название файла пишем число, месяц, фамилию водителя, час выезда
    wb.save('{} {} {} {}.xlsx'.format(txt_chi.get(), txt_mes.get(),
                                      combo_drivers.get(), txt_nac.get()[:2]))

    # Сохраняем словари с водителями и автобусами а данные из формы в
    # файл. Словари меняются редко, но перезаписываем их каждый раз.
    save_settings()


    
#
#Начинаем выполнять код отсюда
#

#
wb = load_workbook('Путевка Шаблон.xlsx')
ws = wb.active

# Создание словарей нужно если файл с настройками
# settings.json пропадёт
##dict_of_buses = {
##    '044':[1,2],
##    '412':[3,4]
##    }
##dict_of_drivers = {
##    'Дакаев А. П.':['Дакаев','Абусаид'],
##    'Дакаев Д. П.':[3,'Давудали']
##    }
##data = {
##    'dict_of_buses': dict_of_buses,
##    'dict_of_drivers': dict_of_drivers,
##    'txt_ser':'',
##    'txt_nom':'',
##    'txt_chi':'',
##    'txt_mes':'',
##    'txt_chiu':'',
##    'txt_mesu':'',
##    'txt_nac':'',
##    'txt_kon':'',
##    'txt_per':''
##    }


# Открываем файл с настройками (settings.json).
# И записываем его содержимое в переменную, а
# словари с автобусами и водителями в свои переменные.
with open("settings.json", "r", encoding="utf-8") as f:
    data = json.load(f)
dict_of_buses = data.get('dict_of_buses')
dict_of_drivers = data.get('dict_of_drivers')

# Создаём окно приложения
window = Tk()  
window.title('Путёвки ООО "ЭКСКОМАВТО"')  
window.geometry('900x350')

# Настраиваем шпифты длч всего окна
default_font = tkFont.nametofont("TkDefaultFont")
default_font.configure(size=14)
default_font = tkFont.nametofont("TkTextFont")
default_font.configure(size=14)

# Создаём три вкладки
tab_control = ttk.Notebook(window)  
tab1 = ttk.Frame(tab_control)  
tab2 = ttk.Frame(tab_control)  
tab3 = ttk.Frame(tab_control)  
tab_control.add(tab1, text='Путёвка')  
tab_control.add(tab2, text='Новый водитель')  
tab_control.add(tab3, text='Новый автобус')  
tab_control.pack(expand=1, fill='both')  

#
# Вкладка "Путёвка"
#

# Размещаем поля для ввода серии и номера путёвки
lb_ser = Label(tab1, text="Серия")  
lb_ser.grid(column=0, row=0)  
txt_ser = Entry(tab1,width=10)  
txt_ser.grid(column=1, row=0)
txt_ser.insert(0,data.get('txt_ser'))

lb_nom = Label(tab1, text="№")  
lb_nom.grid(column=2, row=0)  
txt_nom = Entry(tab1,width=10)  
txt_nom.grid(column=3, row=0)
txt_nom.insert(0,data.get('txt_nom'))

# Размещаем кнопки для увеличения и уменьшения номера путёвки на 1
btn_create = Button(tab1, text="+Номер", command=clic_increment)  
btn_create.grid(column=4, row=0)

btn_create = Button(tab1, text="-Номер", command=clic_decrement)  
btn_create.grid(column=5, row=0)

# Размещаем поля для ввода дат начала и конца смены
lb_chi = Label(tab1, text="Число")  
lb_chi.grid(column=0, row=1)  
txt_chi = Entry(tab1,width=10)  
txt_chi.grid(column=1, row=1)
txt_chi.insert(0,data.get('txt_chi'))

lb_mes = Label(tab1, text="Месяц")  
lb_mes.grid(column=2, row=1)  
txt_mes = Entry(tab1,width=10)  
txt_mes.grid(column=3, row=1)
txt_mes.insert(0,data.get('txt_mes'))

lb_chiu = Label(tab1, text="Число")  
lb_chiu.grid(column=4, row=1)  
txt_chiu = Entry(tab1,width=10)  
txt_chiu.grid(column=5, row=1)
txt_chiu.insert(0,data.get('txt_chiu'))

lb_mesu = Label(tab1, text="Месяц")  
lb_mesu.grid(column=6, row=1)  
txt_mesu = Entry(tab1,width=10)  
txt_mesu.grid(column=7, row=1)
txt_mesu.insert(0,data.get('txt_mesu'))

# Размещаем поля для ввода времени начала и конца смены, а также перерыва
lb_nac = Label(tab1, text="Начало")  
lb_nac.grid(column=0, row=2)  
txt_nac = Entry(tab1,width=10)  
txt_nac.grid(column=1, row=2)
txt_nac.insert(0,data.get('txt_nac'))

lb_kon = Label(tab1, text="Конец")  
lb_kon.grid(column=2, row=2)  
txt_kon = Entry(tab1,width=10)  
txt_kon.grid(column=3, row=2)
txt_kon.insert(0,data.get('txt_kon'))

lb_per = Label(tab1, text="Перерыв")  
lb_per.grid(column=4, row=2)  
txt_per = Entry(tab1,width=15)  
txt_per.grid(column=5, row=2)
txt_per.insert(0,data.get('txt_per'))

# Размещаем комбо-бокс с выбором автобуса
combo_buses = ttk.Combobox(tab1)  
combo_buses['values'] = list(dict_of_buses.keys())
combo_buses.current(0) 
combo_buses.grid(column=0, row=3, columnspan=3)  

# Размещаем комбо-бокс с выбором водителя
combo_drivers = ttk.Combobox(tab1)  
combo_drivers['values'] = list(dict_of_drivers.keys())
combo_drivers.current(0) 
combo_drivers.grid(column=4, row=3, columnspan=3)  

# Размещаем кнопки, заполняющие поля время начала, конца смены и перерыва
btn_create = Button(tab1, text="6-22, 14-17", command=clic_06)  
btn_create.grid(column=0, row=4)
btn_create = Button(tab1, text="7-22, 13-16", command=clic_07)  
btn_create.grid(column=1, row=4)
btn_create = Button(tab1, text="22-7, 1-4", command=clic_22)  
btn_create.grid(column=2, row=4)

btn_create = Button(tab1, text="10-22, 13-16", command=clic_b10)  
btn_create.grid(column=1, row=5)
btn_create = Button(tab1, text="22-10, 1-4", command=clic_b22)  
btn_create.grid(column=2, row=5)

##btn_duble = Button(tab1, text="Число", command=clic_duble)  
##btn_duble.grid(column=3, row=5) 

# Размещаем кнопки увеличения и уменьшения числа начала и конца смены
btn_create = Button(tab1, text="+", command=clic_incdata)  
btn_create.grid(column=3, row=5)

btn_create = Button(tab1, text="-", command=clic_decdata)  
btn_create.grid(column=4, row=5)

btn_create = Button(tab1, text="+", command=clic_incdatau)  
btn_create.grid(column=5, row=5)

btn_create = Button(tab1, text="-", command=clic_decdatau)  
btn_create.grid(column=6, row=5)

# Размещаем кнопку по которой создаётся путёвка, а все поля записываются в файл настроек
btn_create = Button(tab1, text="Создать", command=clic_create)  
btn_create.grid(column=6, row=6)

# Размещаем два комбо-бокса для выбора типов перевозок
combo_price = ttk.Combobox(tab1)
combo_price['values'] = ["Коммерческие перевозки пассажиров и багажа по заказу",
                         "Для собственных нужд"]
combo_price.current(0) 
combo_price.grid(column=3, row=6, columnspan=3)

combo_contry = ttk.Combobox(tab1)  
combo_contry['values'] = ["Городской", "Пригородный", "Межрегиональный"]
combo_contry.current(1) 
combo_contry.grid(column=0, row=6, columnspan=3)


#
# Блок вкладки "Новый водитель"
#

def lbox_drivers_select(event):
    driver_select = lbox_drivers.curselection()
    if len(driver_select) > 0:
        FIO_driver = lbox_drivers.get(driver_select)
        driver = dict_of_drivers.get(FIO_driver)
        txt_FIO.delete(0, END)
        txt_FIO.insert(0, FIO_driver)
        txt_F.delete(0, END)
        txt_F.insert(0, driver[0])
        txt_I.delete(0, END)
        txt_I.insert(0, driver[1])
        txt_O.delete(0, END)
        txt_O.insert(0, driver[2])
        txt_nomer.delete(0, END)
        txt_nomer.insert(0, driver[3])
        txt_rights.delete(0, END)
        txt_rights.insert(0, driver[4])
        txt_card.delete(0, END)
        txt_card.insert(0, driver[5])
        txt_snils.delete(0, END)
        txt_snils.insert(0, driver[6])
        btn_drv_remake["state"] = NORMAL
        btn_drv_delete["state"] = NORMAL
    else:
        btn_drv_delete["state"] = DISABLED





def clic_add_driver():
    driver = []
    driver.append(txt_F.get())
    driver.append(txt_I.get())
    driver.append(txt_O.get())
    driver.append(txt_nomer.get())
    driver.append(txt_rights.get())
    driver.append(txt_card.get())
    driver.append(txt_snils.get())
    dict_of_drivers.update({txt_FIO.get(): driver})
    lbox_drivers.insert(lbox_drivers.size(), txt_FIO.get())
    save_settings()



def clic_remake_driver():
    clic_add_driver()

def clic_delete_driver():
    driver_select = lbox_drivers.curselection()
    dict_of_drivers.pop(lbox_drivers.get(driver_select))
    lbox_drivers.delete(driver_select)
    save_settings()    

# 
lbox_drivers = Listbox(tab2, listvariable=Variable(value=list(dict_of_drivers.keys())), selectmode=SINGLE)
lbox_drivers.grid(column=1, row=1, rowspan=7)
lbox_drivers.bind("<<ListboxSelect>>", lbox_drivers_select)

btn_drv_create = Button(tab2, text="Создать", command=clic_add_driver)  
btn_drv_create.grid(column=3, row=1)

btn_drv_remake = Button(tab2, text="Изменить", command=clic_remake_driver, state=["disabled"])  
btn_drv_remake.grid(column=3, row=3)

btn_drv_delete = Button(tab2, text="Удалить", command=clic_delete_driver, state=["disabled"])  
btn_drv_delete.grid(column=3, row=7)

lb_FIO = Label(tab2, text="ФИО")
lb_FIO.grid(column=5, row=1)
txt_FIO = Entry(tab2, width=10)
txt_FIO.grid(column=6, row=1)

lb_F = Label(tab2, text="Фамилия")
lb_F.grid(column=5, row=3)
txt_F = Entry(tab2, width=10)
txt_F.grid(column=6, row=3)

lb_I = Label(tab2, text="Имя")
lb_I.grid(column=7, row=3)
txt_I = Entry(tab2, width=10)
txt_I.grid(column=8, row=3)

lb_O = Label(tab2, text="Отчество")
lb_O.grid(column=5, row=4)
txt_O = Entry(tab2, width=10)
txt_O.grid(column=6, row=4)

lb_nomer = Label(tab2, text="Номер")
lb_nomer.grid(column=5, row=5)
txt_nomer = Entry(tab2, width=10)
txt_nomer.grid(column=6, row=5)

lb_rights = Label(tab2, text="Права")
lb_rights.grid(column=7, row=5)
txt_rights = Entry(tab2, width=10)
txt_rights.grid(column=8, row=5)

lb_card = Label(tab2, text="Карта")
lb_card.grid(column=5, row=6)
txt_card = Entry(tab2, width=10)
txt_card.grid(column=6, row=6)

lb_snils = Label(tab2, text="СНИЛС")
lb_snils.grid(column=7, row=6)
txt_snils = Entry(tab2, width=10)
txt_snils.grid(column=8, row=6)


#
# Блок вкладки "Новый водитель"
#

def lbox_buses_select(event):
    bus_select = lbox_buses.curselection()
    if len(driver_select) > 0:
        FIO_driver = lbox_buses.get(bus_select)
        bus = dict_of_buses.get(FIO_driver)
        txt_gar_nomer.delete(0, END)
        txt_gar_nomer.insert(0, FIO_driver)
        txt_F.delete(0, END)
        txt_F.insert(0, bus[0])
        txt_I.delete(0, END)
        txt_I.insert(0, bus[1])
        txt_O.delete(0, END)
        txt_O.insert(0, bus[2])
        txt_nomer.delete(0, END)
        txt_nomer.insert(0, bus[3])
        txt_rights.delete(0, END)
        txt_rights.insert(0, bus[4])
        txt_card.delete(0, END)
        txt_card.insert(0, bus[5])
        txt_snils.delete(0, END)
        txt_snils.insert(0, bus[6])
        btn_bus_remake["state"] = NORMAL
        btn_bus_delete["state"] = NORMAL
    else:
        btn_bus_delete["state"] = DISABLED





def clic_add_bus():
    bus = []
    bus.append(txt_F.get())
    bus.append(txt_I.get())
    bus.append(txt_O.get())
    bus.append(txt_nomer.get())
    bus.append(txt_rights.get())
    bus.append(txt_card.get())
    bus.append(txt_snils.get())
    dict_of_buses.update({txt_gar_nomer.get(): bus})
    lbox_buses.insert(lbox_buses.size(), txt_gar_nomer.get())
    save_settings()



def clic_remake_bus():
    clic_add_bus()

def clic_delete_bus():
    bus_select = lbox_buses.curselection()
    dict_of_buses.pop(lbox_buses.get(bus_select))
    lbox_buses.delete(bus_select)
    save_settings()    

# 
lbox_buses = Listbox(tab3, listvariable=Variable(value=list(dict_of_drivers.keys())), selectmode=SINGLE)
lbox_buses.grid(column=1, row=1, rowspan=7)
lbox_buses.bind("<<ListboxSelect>>", lbox_buses_select)

btn_bus_create = Button(tab3, text="Создать", command=clic_add_bus)  
btn_bus_create.grid(column=3, row=1)

btn_bus_remake = Button(tab3, text="Изменить", command=clic_remake_bus, state=["disabled"])  
btn_bus_remake.grid(column=3, row=3)

btn_bus_delete = Button(tab3, text="Удалить", command=clic_delete_bus, state=["disabled"])  
btn_bus_delete.grid(column=3, row=7)

lb_gar_nomer = Label(tab3, text="Гаражный номер")
lb_gar_nomer.grid(column=5, row=1)
txt_gar_nomer = Entry(tab3, width=10)
txt_gar_nomer.grid(column=6, row=1)

lb_gos_nomer = Label(tab3, text="Гос. номер")
lb_gos_nomer.grid(column=5, row=3)
txt_gos_nomer = Entry(tab3, width=10)
txt_gos_nomer.grid(column=6, row=3)

lb_cat = Label(tab3, text="Категория")
lb_cat.grid(column=7, row=3)
txt_cat = Entry(tab3, width=10)
txt_cat.grid(column=8, row=3)

lb_make = Label(tab3, text="Марка")
lb_make.grid(column=5, row=4)
txt_make = Entry(tab3, width=10)
txt_make.grid(column=6, row=4)

lb_model = Label(tab3, text="Модель")
lb_model.grid(column=7, row=4)
txt_model = Entry(tab3, width=10)
txt_model.grid(column=8, row=4)

lb_owner = Label(tab3, text="Владелец")
lb_owner.grid(column=7, row=4)
txt_owner = Entry(tab3, width=10)
txt_owner.grid(column=8, row=4)

lb_master = Label(tab3, text="Выпускающая организация (полное название)")
lb_master.grid(column=7, row=4)
txt_master = Entry(tab3, width=10)
txt_master.grid(column=8, row=4)

# Включаем обработчик событий окна
window.mainloop()
