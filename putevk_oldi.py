from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import ttk
import tkinter.font as tkFont
import json
import time


def clic_06():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "06:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "22:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 14:00 до 17:00")
    
def clic_07():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "07:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "22:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 13:00 до 16:00")

def clic_22():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "22:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "07:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 01:00 до 04:00")

def clic_b10():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "10:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "22:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 13:00 до 16:00")

def clic_b22():
    txt_nac.delete(0, END)
    txt_nac.insert(0, "22:00")
    txt_kon.delete(0, END)
    txt_kon.insert(0, "10:00")
    txt_per.delete(0, END)
    txt_per.insert(0, "с 01:00 до 04:00")
    
def clic_duble():
    txt_chiu.delete(0, END)
    txt_chiu.insert(0,txt_chi.get())
    txt_mesu.delete(0, END)
    txt_mesu.insert(0,txt_mes.get())

def clic_increment():
    a = str(int(txt_nom.get()) + 1)
    if len(a)==1: a='0'+a
    txt_nom.delete(0, END)
    txt_nom.insert(0,a)

def clic_decrement():
    a = str(int(txt_nom.get()) - 1)
    if len(a)==1: a='0'+a
    txt_nom.delete(0, END)
    txt_nom.insert(0,a)

def clic_incdata():
    a = str(int(txt_chi.get()) + 1)
    if len(a)==1: a='0'+a
    txt_chi.delete(0, END)
    txt_chi.insert(0,a)

def clic_decdata():
    a = str(int(txt_chi.get()) - 1)
    if len(a)==1: a='0'+a
    txt_chi.delete(0, END)
    txt_chi.insert(0,a)

def clic_incdatau():
    a = str(int(txt_chiu.get()) + 1)
    if len(a)==1: a='0'+a
    txt_chiu.delete(0, END)
    txt_chiu.insert(0,a)

def clic_decdatau():
    a = str(int(txt_chiu.get()) - 1)
    if len(a)==1: a='0'+a
    txt_chiu.delete(0, END)
    txt_chiu.insert(0,a)


def clic_create():
    ws['BZ2'] = txt_ser.get()
    ws['CJ2'] = txt_nom.get()
    ws['AN4'] = txt_chi.get()
    ws['AS4'] = txt_mes.get()
    ws['BY4'] = txt_chiu.get()
    ws['CG4'] = txt_mesu.get()
    ws['BQ6'] = dict_of_buses.get(combo_buses.get())[0]
    ws['CZ6'] = dict_of_buses.get(combo_buses.get())[1]
    ws['CH7'] = dict_of_buses.get(combo_buses.get())[3]
    ws['CC9'] = dict_of_buses.get(combo_buses.get())[4]
    ws['BY16'] = txt_nac.get()
    ws['BY18'] = txt_kon.get()
    ws['BY20'] = txt_per.get()
    ws['P25'] = combo_drivers.get()
    ws['P51'] = combo_drivers.get()
    ws['BJ30'] = dict_of_drivers.get(combo_drivers.get())[0]
    ws['BJ32'] = dict_of_drivers.get(combo_drivers.get())[1]
    ws['BJ34'] = dict_of_drivers.get(combo_drivers.get())[2]
    ws['CQ30'] = dict_of_drivers.get(combo_drivers.get())[3]
    ws['DC30'] = dict_of_drivers.get(combo_drivers.get())[4]
    ws['DN30'] = dict_of_drivers.get(combo_drivers.get())[5]
    ws['BZ10'] = combo_price.get()
    ws['BQ11'] = combo_contry.get()
    ws.title = '{} {}'.format(combo_drivers.get(), txt_nac.get()[:2])


    wb.save('{} {} {} {}.xlsx'.format(txt_chi.get(), txt_mes.get(), combo_drivers.get(), txt_nac.get()[:2]))
    
    data = {
        'dict_of_buses': dict_of_buses,
        'dict_of_drivers': dict_of_drivers,
        'txt_ser':txt_ser.get(),
        'txt_nom':txt_nom.get(),
        'txt_chi':txt_chi.get(),
        'txt_mes':txt_mes.get(),
        'txt_chiu':txt_chiu.get(),
        'txt_mesu':txt_mesu.get(),
        'txt_nac':txt_nac.get(),
        'txt_kon':txt_kon.get(),
        'txt_per':txt_per.get()
        }
    with open("settings.json", "w", encoding="utf-8") as write_file:
        json.dump(data, write_file, ensure_ascii=False, indent='  ')
    


wb = load_workbook('Путевка Шаблон.xlsx')
ws = wb.active


dict_of_buses = {
    '044':[1,2],
    '412':[3,4]
    }

dict_of_drivers = {
    'Дакаев А. П.':['Дакаев','Абусаид'],
    'Дакаев Д. П.':[3,'Давудали']
    }
with open("settings.json", "r", encoding="utf-8") as f:
    data = json.load(f)
dict_of_buses = data.get('dict_of_buses')
dict_of_drivers = data.get('dict_of_drivers')


window = Tk()  
window.title('Путёвки ООО "ЭКСКОМАВТО"')  
window.geometry('900x350')

default_font = tkFont.nametofont("TkDefaultFont")
default_font.configure(size=14)
default_font = tkFont.nametofont("TkTextFont")
default_font.configure(size=14)

tab_control = ttk.Notebook(window)  
tab1 = ttk.Frame(tab_control)  
tab2 = ttk.Frame(tab_control)  
tab3 = ttk.Frame(tab_control)  
tab_control.add(tab1, text='Путёвка')  
tab_control.add(tab2, text='Новый автобус')  
tab_control.add(tab3, text='Новый водтель')  
tab_control.pack(expand=1, fill='both')  


lb_ser = Label(tab1, text="Серия")  
lb_ser.grid(column=0, row=0)  
txt_ser = Entry(tab1,width=10)  
txt_ser.grid(column=1, row=0)
txt_ser.insert(0,data.get('txt_ser'))

lb_nom = Label(tab1, text="№")  
lb_nom.grid(column=2, row=0)  
txt_nom = Entry(tab1,width=10)  
txt_nom.grid(column=3, row=0)
txt_nom.insert(0,data.get('txt_nom'))

btn_create = Button(tab1, text="+Номер", command=clic_increment)  
btn_create.grid(column=4, row=0)

btn_create = Button(tab1, text="-Номер", command=clic_decrement)  
btn_create.grid(column=5, row=0)

lb_chi = Label(tab1, text="Число")  
lb_chi.grid(column=0, row=1)  
txt_chi = Entry(tab1,width=10)  
txt_chi.grid(column=1, row=1)
txt_chi.insert(0,data.get('txt_chi'))

lb_mes = Label(tab1, text="Месяц")  
lb_mes.grid(column=2, row=1)  
txt_mes = Entry(tab1,width=10)  
txt_mes.grid(column=3, row=1)
txt_mes.insert(0,data.get('txt_mes'))

lb_chiu = Label(tab1, text="Число")  
lb_chiu.grid(column=4, row=1)  
txt_chiu = Entry(tab1,width=10)  
txt_chiu.grid(column=5, row=1)
txt_chiu.insert(0,data.get('txt_chiu'))

lb_mesu = Label(tab1, text="Месяц")  
lb_mesu.grid(column=6, row=1)  
txt_mesu = Entry(tab1,width=10)  
txt_mesu.grid(column=7, row=1)
txt_mesu.insert(0,data.get('txt_mesu'))


lb_nac = Label(tab1, text="Начало")  
lb_nac.grid(column=0, row=2)  
txt_nac = Entry(tab1,width=10)  
txt_nac.grid(column=1, row=2)
txt_nac.insert(0,data.get('txt_nac'))

lb_kon = Label(tab1, text="Конец")  
lb_kon.grid(column=2, row=2)  
txt_kon = Entry(tab1,width=10)  
txt_kon.grid(column=3, row=2)
txt_kon.insert(0,data.get('txt_kon'))

lb_per = Label(tab1, text="Перерыв")  
lb_per.grid(column=4, row=2)  
txt_per = Entry(tab1,width=15)  
txt_per.grid(column=5, row=2)
txt_per.insert(0,data.get('txt_per'))

combo_buses = ttk.Combobox(tab1)  
combo_buses['values'] = list(dict_of_buses.keys())
combo_buses.current(0) 
combo_buses.grid(column=0, row=3, columnspan=3)  

combo_drivers = ttk.Combobox(tab1)  
combo_drivers['values'] = list(dict_of_drivers.keys())
combo_drivers.current(0) 
combo_drivers.grid(column=4, row=3, columnspan=3)  

btn_create = Button(tab1, text="6-22, 14-17", command=clic_06)  
btn_create.grid(column=0, row=4)
btn_create = Button(tab1, text="7-22, 13-16", command=clic_07)  
btn_create.grid(column=1, row=4)
btn_create = Button(tab1, text="22-7, 1-4", command=clic_22)  
btn_create.grid(column=2, row=4)

btn_create = Button(tab1, text="10-22, 13-16", command=clic_b10)  
btn_create.grid(column=1, row=5)
btn_create = Button(tab1, text="22-10, 1-4", command=clic_b22)  
btn_create.grid(column=2, row=5)

##btn_duble = Button(tab1, text="Число", command=clic_duble)  
##btn_duble.grid(column=3, row=5) 


btn_create = Button(tab1, text="+", command=clic_incdata)  
btn_create.grid(column=3, row=5)

btn_create = Button(tab1, text="-", command=clic_decdata)  
btn_create.grid(column=4, row=5)

btn_create = Button(tab1, text="+", command=clic_incdatau)  
btn_create.grid(column=5, row=5)

btn_create = Button(tab1, text="-", command=clic_decdatau)  
btn_create.grid(column=6, row=5)


btn_create = Button(tab1, text="Создать", command=clic_create)  
btn_create.grid(column=6, row=6)


 
combo_price = ttk.Combobox(tab1)
combo_price['values'] = ["Коммерческие перевозки пассажиров и багажа по заказу",
                         "Для собственных нужд"]
combo_price.current(0) 
combo_price.grid(column=3, row=6, columnspan=3)

combo_contry = ttk.Combobox(tab1)  
combo_contry['values'] = ["Городской", "Пригородный", "Межрегиональный"]
combo_contry.current(1) 
combo_contry.grid(column=0, row=6, columnspan=3)

window.mainloop()



